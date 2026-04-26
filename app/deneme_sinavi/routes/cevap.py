"""Toplu cevap girisi (sube × ogrenci × ders grid'i) + Excel import/export."""
import io
from datetime import datetime

from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, send_file, abort)
from flask_login import login_required

from app.utils import role_required
from app.extensions import db
from app.models.deneme_sinavi import (DenemeSinavi, DenemeDersi,
                                      DenemeKatilim, DenemeDersSonucu)
from app.models.kayit import Sube, OgrenciKayit
from app.deneme_sinavi.hesaplar import guncelle_katilim_toplamlari
from app.deneme_sinavi.bildirim import bildirim_gonder_katilimlar


bp = Blueprint('cevap', __name__, url_prefix='/sinav/<int:sinav_id>/cevap')


def _sube_ogrencileri(sube_id):
    """Subedeki aktif ogrencileri dondur (Ogrenci nesnesi listesi)."""
    kayitlar = (OgrenciKayit.query
                .filter_by(sube_id=sube_id, durum='aktif')
                .all())
    return [k.ogrenci for k in kayitlar if k.ogrenci and k.ogrenci.aktif]


def _get_or_create_katilim(sinav_id, ogrenci_id, sube_id):
    k = DenemeKatilim.query.filter_by(
        deneme_sinavi_id=sinav_id,
        ogrenci_id=ogrenci_id,
    ).first()
    if k:
        if sube_id and not k.sube_id:
            k.sube_id = sube_id
        return k
    k = DenemeKatilim(
        deneme_sinavi_id=sinav_id,
        ogrenci_id=ogrenci_id,
        sube_id=sube_id,
        katildi=True,
    )
    db.session.add(k)
    db.session.flush()
    return k


def _get_or_create_ders_sonucu(katilim_id, ders_id):
    s = DenemeDersSonucu.query.filter_by(
        katilim_id=katilim_id,
        deneme_dersi_id=ders_id,
    ).first()
    if s:
        return s
    s = DenemeDersSonucu(
        katilim_id=katilim_id,
        deneme_dersi_id=ders_id,
        dogru=0, yanlis=0, bos=0, net=0,
    )
    db.session.add(s)
    db.session.flush()
    return s


@bp.route('/', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'ogretmen')
def giris(sinav_id):
    """Sube sec + grid'e cevap gir."""
    from app.deneme_sinavi.kategori import (sinav_tipi_kategorisi,
                                             ogrenci_uygun_mu,
                                             ogrenci_kategorisi)
    sinav = DenemeSinavi.query.get_or_404(sinav_id)
    dersler = sinav.dersler.all()
    subeler = Sube.query.filter_by(aktif=True).order_by(Sube.ad).all()

    sube_id = request.args.get('sube_id', type=int) or request.form.get('sube_id', type=int)
    # Kategori filtresi gevsetme bayragi: tum_ogrenciler=1 → kategori dışındakileri de göster
    tum_ogrenciler = (request.args.get('tum_ogrenciler') == '1') or \
                     (request.form.get('tum_ogrenciler') == '1')
    sinav_kategorisi = sinav_tipi_kategorisi(sinav.sinav_tipi)

    ogrenci_satirlari = []
    filtrelenen_sayisi = 0  # Kategori uymadigi icin gizlenen ogrenci adedi
    if sube_id:
        ogrenciler = _sube_ogrencileri(sube_id)
        for o in ogrenciler:
            # Kategori filtresi: LGS sınavda LGS-uygun (6-8) öğrenciler,
            # YKS sınavda YKS-uygun (9-12, TYT/AYT/Mezun) öğrenciler.
            if not tum_ogrenciler:
                if not ogrenci_uygun_mu(o, sinav_kategorisi, kategorisiz_dahil=True):
                    filtrelenen_sayisi += 1
                    continue
            katilim = DenemeKatilim.query.filter_by(
                deneme_sinavi_id=sinav.id, ogrenci_id=o.id
            ).first()
            # Mevcut sonuclari ders bazinda lookup'la
            mevcut = {}
            if katilim:
                for s in katilim.ders_sonuclari:
                    mevcut[s.deneme_dersi_id] = s
            ogrenci_satirlari.append({
                'ogrenci': o,
                'katilim': katilim,
                'sonuclar': mevcut,
                'og_kategori': ogrenci_kategorisi(o),
            })

    if request.method == 'POST' and sube_id:
        # Form alan ismi: d_{ogrenci_id}_{ders_id}_{D|Y|B}
        kaydedilen = 0
        bildirim_katilimlari: list[DenemeKatilim] = []
        bildirim_gonder = request.form.get('bildirim_gonder') == '1'
        for row in ogrenci_satirlari:
            o = row['ogrenci']
            herhangi_girildi = False
            bekleyen_sonuclar = []
            for ders in dersler:
                key_d = f'd_{o.id}_{ders.id}_D'
                key_y = f'd_{o.id}_{ders.id}_Y'
                key_b = f'd_{o.id}_{ders.id}_B'
                try:
                    d = int(request.form.get(key_d) or 0)
                    y = int(request.form.get(key_y) or 0)
                    b = int(request.form.get(key_b) or 0)
                except (TypeError, ValueError):
                    continue
                if d or y or b:
                    herhangi_girildi = True
                # D+Y+B soru_sayisindan buyukse sinirla
                if (d + y + b) > (ders.soru_sayisi or 0) and ders.soru_sayisi:
                    # kullaniciya feedback vermek yerine sessizce kirpma yok, olduğu gibi kaydet
                    pass
                bekleyen_sonuclar.append((ders, d, y, b))

            if not herhangi_girildi:
                # Bu ogrenci icin hicbir sey girilmediyse eski kayitlari dokunma
                continue

            # OBP (LGS icin opsiyonel)
            obp_key = f'obp_{o.id}'
            obp_str = request.form.get(obp_key, '').strip()
            obp = None
            if obp_str:
                try:
                    obp = float(obp_str)
                except ValueError:
                    obp = None

            katilim = _get_or_create_katilim(sinav.id, o.id, sube_id)
            katilim.obp = obp
            katilim.katildi = True
            for ders, d, y, b in bekleyen_sonuclar:
                s = _get_or_create_ders_sonucu(katilim.id, ders.id)
                s.dogru, s.yanlis, s.bos = d, y, b
                s.hesapla_net()
            guncelle_katilim_toplamlari(katilim)
            kaydedilen += 1
            bildirim_katilimlari.append(katilim)

        db.session.commit()

        bildirim_sayisi = 0
        if bildirim_gonder and bildirim_katilimlari:
            bildirim_sayisi = bildirim_gonder_katilimlar(bildirim_katilimlari)
            db.session.commit()

        msg = f'{kaydedilen} ogrencinin sonuclari kaydedildi.'
        if bildirim_sayisi:
            msg += f' {bildirim_sayisi} bildirim gonderildi.'
        flash(msg, 'success')
        return redirect(url_for('deneme_sinavi.cevap.giris',
                                sinav_id=sinav.id, sube_id=sube_id))

    return render_template('deneme_sinavi/cevap_giris.html',
                           sinav=sinav,
                           dersler=dersler,
                           subeler=subeler,
                           sube_id=sube_id,
                           ogrenci_satirlari=ogrenci_satirlari,
                           sinav_kategorisi=sinav_kategorisi,
                           tum_ogrenciler=tum_ogrenciler,
                           filtrelenen_sayisi=filtrelenen_sayisi)


# === Excel Import/Export ===

def _sablon_baslik_satiri(dersler):
    """Excel sablon basligi."""
    basliklar = ['Ogrenci No', 'Ad Soyad', 'Sube']
    for d in dersler:
        basliklar.extend([f'{d.ders_adi}_D',
                          f'{d.ders_adi}_Y',
                          f'{d.ders_adi}_B'])
    basliklar.append('OBP (opsiyonel)')
    return basliklar


@bp.route('/sablon')
@login_required
@role_required('admin', 'ogretmen')
def sablon(sinav_id):
    """Excel sablonunu indir (subedeki ogrenciler pre-fill)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sinav = DenemeSinavi.query.get_or_404(sinav_id)
    dersler = sinav.dersler.all()
    sube_id = request.args.get('sube_id', type=int)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cevap Girisi'

    # Baslik
    basliklar = _sablon_baslik_satiri(dersler)
    for col, h in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F46E5')

    # Ogrenci satirlari (opsiyonel pre-fill)
    if sube_id:
        sube = Sube.query.get(sube_id)
        sube_ad = sube.tam_ad if sube else ''
        ogrenciler = _sube_ogrencileri(sube_id)
        for row_idx, o in enumerate(ogrenciler, 2):
            ws.cell(row=row_idx, column=1, value=o.ogrenci_no)
            ws.cell(row=row_idx, column=2, value=o.tam_ad)
            ws.cell(row=row_idx, column=3, value=sube_ad)

    # Sutun genislikleri
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"deneme_{sinav.id}_sablon_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/import', methods=['POST'])
@login_required
@role_required('admin', 'ogretmen')
def excel_import(sinav_id):
    """Excel dosyasindan toplu cevap import'u."""
    from openpyxl import load_workbook
    from app.models.muhasebe import Ogrenci

    sinav = DenemeSinavi.query.get_or_404(sinav_id)
    dersler = sinav.dersler.all()
    ders_by_ad = {d.ders_adi: d for d in dersler}

    if 'dosya' not in request.files or request.files['dosya'].filename == '':
        flash('Dosya secilmedi.', 'warning')
        return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))

    dosya = request.files['dosya']
    try:
        wb = load_workbook(dosya, read_only=True, data_only=True)
    except Exception as e:
        flash(f'Dosya acilamadi: {e}', 'danger')
        return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))

    ws = wb.active
    # Basligi oku
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash('Bos dosya.', 'warning')
        return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))
    baslik = [str(x) if x is not None else '' for x in rows[0]]

    # Sutun indekslerini cikar
    try:
        ogr_no_col = baslik.index('Ogrenci No')
    except ValueError:
        flash('Basliklar eslesmiyor: "Ogrenci No" sutunu yok.', 'danger')
        return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))

    obp_col = None
    if 'OBP (opsiyonel)' in baslik:
        obp_col = baslik.index('OBP (opsiyonel)')

    # Her ders icin D/Y/B sutunlari
    ders_sutun = {}  # ders -> (d_col, y_col, b_col)
    for d in dersler:
        try:
            d_col = baslik.index(f'{d.ders_adi}_D')
            y_col = baslik.index(f'{d.ders_adi}_Y')
            b_col = baslik.index(f'{d.ders_adi}_B')
            ders_sutun[d.id] = (d_col, y_col, b_col)
        except ValueError:
            # Bu ders icin sutun yok, atla
            continue

    if not ders_sutun:
        flash('Dosyada hicbir ders sutunu bulunamadi.', 'danger')
        return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))

    bildirim_gonder = request.form.get('bildirim_gonder') == '1'
    bildirim_katilimlari: list[DenemeKatilim] = []

    eklenen = 0
    atlanan = 0
    for row in rows[1:]:
        if not row or len(row) <= ogr_no_col:
            continue
        ogr_no = row[ogr_no_col]
        if not ogr_no:
            continue
        ogr = Ogrenci.query.filter_by(ogrenci_no=str(ogr_no).strip()).first()
        if not ogr:
            atlanan += 1
            continue

        # Ogrencinin aktif kaydindaki sube
        aktif_k = OgrenciKayit.query.filter_by(
            ogrenci_id=ogr.id, durum='aktif').first()
        sube_id = aktif_k.sube_id if aktif_k else None

        katilim = _get_or_create_katilim(sinav.id, ogr.id, sube_id)
        katilim.katildi = True

        # OBP
        if obp_col is not None and len(row) > obp_col and row[obp_col] is not None:
            try:
                katilim.obp = float(row[obp_col])
            except (ValueError, TypeError):
                pass

        for ders in dersler:
            if ders.id not in ders_sutun:
                continue
            d_col, y_col, b_col = ders_sutun[ders.id]
            try:
                d = int(row[d_col]) if len(row) > d_col and row[d_col] is not None else 0
                y = int(row[y_col]) if len(row) > y_col and row[y_col] is not None else 0
                b = int(row[b_col]) if len(row) > b_col and row[b_col] is not None else 0
            except (ValueError, TypeError):
                continue
            s = _get_or_create_ders_sonucu(katilim.id, ders.id)
            s.dogru, s.yanlis, s.bos = d, y, b
            s.hesapla_net()

        guncelle_katilim_toplamlari(katilim)
        eklenen += 1
        bildirim_katilimlari.append(katilim)

    db.session.commit()

    bildirim_sayisi = 0
    if bildirim_gonder and bildirim_katilimlari:
        bildirim_sayisi = bildirim_gonder_katilimlar(bildirim_katilimlari)
        db.session.commit()

    msg = f'{eklenen} ogrencinin sonuclari import edildi.'
    if atlanan:
        msg += f' {atlanan} satir ogrenci no eslesmedigi icin atlandi.'
    if bildirim_sayisi:
        msg += f' {bildirim_sayisi} bildirim gonderildi.'
    flash(msg, 'success')
    return redirect(url_for('deneme_sinavi.cevap.giris', sinav_id=sinav.id))
