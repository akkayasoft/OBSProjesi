"""Toplu yukleme (Excel import) yardimci fonksiyonlari."""
from datetime import date, datetime
from io import BytesIO
from decimal import Decimal, InvalidOperation


def excel_sablonu_olustur(basliklar, ornek_satirlar=None, aciklama_satiri=None, sayfa_adi='Veri'):
    """Verilen basliklarla Excel sablonu olusturur.

    basliklar: list[dict] — her biri {'key': 'ogrenci_no', 'label': 'Öğrenci No', 'zorunlu': True, 'ornek': '1001'}
    ornek_satirlar: list[list] — basliklar sirasina uygun ornek veri satirlari
    aciklama_satiri: ilk satira yazilacak kullanici aciklamasi (istege bagli)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sayfa_adi

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    baslangic_satir = 1
    if aciklama_satiri:
        ws.cell(row=1, column=1, value=aciklama_satiri)
        ws.cell(row=1, column=1).font = Font(italic=True, color='666666', size=10)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(basliklar))
        baslangic_satir = 2

    # Basliklar
    header_fill = PatternFill('solid', fgColor='2F6DB5')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    zorunlu_fill = PatternFill('solid', fgColor='D93A3A')

    for idx, b in enumerate(basliklar, start=1):
        cell = ws.cell(row=baslangic_satir, column=idx, value=b['label'] + (' *' if b.get('zorunlu') else ''))
        cell.font = header_font
        cell.fill = header_fill if not b.get('zorunlu') else zorunlu_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        # Kolon genisligi
        ws.column_dimensions[get_column_letter(idx)].width = max(len(b['label']) + 4, 15)

    ws.row_dimensions[baslangic_satir].height = 30

    # Ornek satirlar
    if ornek_satirlar:
        for r_idx, satir in enumerate(ornek_satirlar, start=baslangic_satir + 1):
            for c_idx, val in enumerate(satir, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.border = border
                c.font = Font(italic=True, color='888888')

    # Donen: BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _normalize_header(s):
    """Baslik karsilastirmasi icin normalize et: kucuk harf, bosluksuz,
    '*' ve parantezli notlari at. 'Öğrenci No *' == 'ogrenci no'."""
    import re as _re
    s = str(s or '').strip()
    # '(erkek/kadin)', '(YYYY-MM-DD)' gibi aciklamalari kaldir
    s = _re.sub(r'\([^)]*\)', '', s)
    s = s.replace('*', '').strip().lower()
    # Turkce harf normalize
    tr_map = {'ı': 'i', 'ş': 's', 'ğ': 'g', 'ü': 'u', 'ö': 'o', 'ç': 'c'}
    for a, b in tr_map.items():
        s = s.replace(a, b)
    # Bosluklari ve noktalama kaldir
    s = _re.sub(r'[\s\-_.]+', '', s)
    return s


def excel_oku(dosya, basliklar, max_header_satir=5):
    """Yuklenen Excel dosyasini parse eder.

    basliklar: excel_sablonu_olustur ile ayni format
                (list[dict] -- {'key','label','zorunlu'})
    max_header_satir: basligin aranacagi ilk N satir (aciklama satiri, bos
                satir vs. dusunulerek biraz esnek).

    Veriyi kolonun POZISYONU yerine BASLIK ADIYLA eslestirir — yani
    kullanici Excel'de kolon sirasini degistirse ya da araya kolon eklese
    bile dogru degerler ilgili alanlara gider. Bilinmeyen kolonlar atlanir,
    bulunamayan ZORUNLU basliklar icin net bir hata firlatilir.

    Donus: list[dict] — her biri key -> deger map'i, ek olarak
           '_satir_no' ve '_hatalar' alanlari.
    """
    from openpyxl import load_workbook

    wb = load_workbook(dosya, data_only=True)
    ws = wb.active

    # Beklenen basliklari normalize et (label VE key ikisine de bakacagiz)
    beklenen = {}
    for b in basliklar:
        beklenen[_normalize_header(b['label'])] = b['key']
        beklenen[_normalize_header(b['key'])] = b['key']

    # Baslik satirini ara: en cok eslesme veren satir. Tum kolonlari tarariz
    # (100 kolona kadar yeter).
    max_col = min(ws.max_column or len(basliklar) * 2, 100)
    baslik_satir_no = None
    col_map = {}  # key -> excel col idx
    en_iyi_eslesme = 0

    for r in range(1, max_header_satir + 1):
        guncel_map = {}
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            norm = _normalize_header(raw)
            if norm in beklenen:
                key = beklenen[norm]
                # Ayni anahtarin iki kolonu varsa ilk bulunan kullanilir
                if key not in guncel_map:
                    guncel_map[key] = c
        if len(guncel_map) > en_iyi_eslesme:
            en_iyi_eslesme = len(guncel_map)
            baslik_satir_no = r
            col_map = guncel_map

    # En az zorunlu basliklarin tamami bulunmali
    zorunlu_keys = [b['key'] for b in basliklar if b.get('zorunlu')]
    eksik_zorunlu = [b['label'] for b in basliklar
                     if b.get('zorunlu') and b['key'] not in col_map]

    if baslik_satir_no is None or en_iyi_eslesme < min(3, len(basliklar)):
        raise ValueError(
            'Şablon başlıkları bulunamadı. Lütfen indirilen şablonu kullanın '
            've başlık satırını değiştirmeyin.'
        )

    if eksik_zorunlu:
        raise ValueError(
            'Excel\'de şu zorunlu sütun(lar) bulunamadı: '
            + ', '.join(eksik_zorunlu)
            + '. Şablonu yeniden indirip başlıkları değiştirmeden doldurun.'
        )

    satirlar = []
    for r in range(baslik_satir_no + 1, (ws.max_row or baslik_satir_no) + 1):
        satir = {'_satir_no': r, '_hatalar': []}
        hepsi_bos = True

        # Kolon-eslemesi uzerinden oku (pozisyon degil, baslik-esli)
        for key, col_idx in col_map.items():
            val = ws.cell(row=r, column=col_idx).value
            if val is not None and str(val).strip() != '':
                hepsi_bos = False
            satir[key] = val

        # Bilinmeyen / bulunmayan basliklari None yap
        for b in basliklar:
            satir.setdefault(b['key'], None)

        if hepsi_bos:
            continue

        satirlar.append(satir)

    return satirlar


def dogrula_str(val, zorunlu=False, max_len=None, label=''):
    """String degeri dogrular, hata listesi doner."""
    if val is None or str(val).strip() == '':
        if zorunlu:
            return None, f'{label} zorunludur.'
        return None, None
    s = str(val).strip()
    if max_len and len(s) > max_len:
        return None, f'{label} en fazla {max_len} karakter olabilir.'
    return s, None


def dogrula_sayi(val, zorunlu=False, min_val=None, max_val=None, tam_sayi=False, label=''):
    if val is None or str(val).strip() == '':
        if zorunlu:
            return None, f'{label} zorunludur.'
        return None, None
    try:
        if tam_sayi:
            n = int(float(str(val).replace(',', '.').strip()))
        else:
            n = Decimal(str(val).replace(',', '.').strip())
    except (ValueError, InvalidOperation):
        return None, f'{label} geçerli bir sayı olmalı.'

    if min_val is not None and n < min_val:
        return None, f'{label} en az {min_val} olmalı.'
    if max_val is not None and n > max_val:
        return None, f'{label} en fazla {max_val} olabilir.'
    return n, None


def dogrula_tarih(val, zorunlu=False, label=''):
    if val is None or (isinstance(val, str) and val.strip() == ''):
        if zorunlu:
            return None, f'{label} zorunludur.'
        return None, None
    if isinstance(val, datetime):
        return val.date(), None
    if isinstance(val, date):
        return val, None
    # String olarak gelmisse parse
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date(), None
        except ValueError:
            continue
    return None, f'{label} gecerli bir tarih olmalı (YYYY-MM-DD veya DD.MM.YYYY).'
